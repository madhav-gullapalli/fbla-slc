import kivy
#run app
from kivy.app import App
# ui elements
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.uix.slider import Slider
#screens
from kivy.uix.screenmanager import ScreenManager, Screen
#code that creates static buttons
from kivy.lang import Builder
#setSize
from kivy.core.window import Window
#set slider value
from kivy.properties  import NumericProperty
#use excel
import openpyxl
#use web
import webbrowser

Window.size=(800,600)               
#file to load static ui elements(ones that just display singular value)
#dynamic ui elements are put in classes ( ones that display dynamic data)
Builder.load_string("""
#:kivy 1.11.0
#the starting 'menu' screen
<startScreen>:
    BoxLayout:
        Label:
            text:'to skip click skip button'
        MainWidget:
            size_hint_x:1
        


#filtered based on type
<typesScreen>:
    BoxLayout:
        orientation: 'vertical'
        Label:
            text:" what type of attraction do you want to go "
            size_hint_y: 0.5
        Label:
            text:"( just type in text in [])"
            size_hint_y: 0.5
        Label:
            text:"[restaurant],[zoos/theme parks],[museums],singular significant [landmark],[areas] with multiple landmarks"
            size_hint_y:0.5
        TypeInput:
#filtered based on area
<areasScreen>:
    BoxLayout:
        orientation: 'vertical'
        Label:
            text:" pick what region you want to go to?"
            size_hint_y: 0.5
        Label:
            text:"[atlanta,eastern,coast,mountains,central,southern]"
            size_hint_y: 0.5
        Label:
            text:""
            size_hint_y:0.5
        areaInput:  
#filtered based on surrounding  
<SurroundScreen>:
    BoxLayout:
        orientation: 'vertical'
        Label:
            text:"Do you want to be in an [urban],[suburban], or [rural] area?"
            size_hint_y: 0.5
        Label:
            text:""
            size_hint_y: 0.5
        Label:
            text:""
            size_hint_y:0.5
        surroundInput:   
#filtered based on wheter you want outside attraction
<OutsideScreen>:
    BoxLayout:
        orientation: 'vertical'
        Label:
            text:"Do you want to go to a outside location? [yes] or [no]"
            size_hint_y: 0.5
        Label:
            text:""
            size_hint_y: 0.5
        Label:
            text:""
            size_hint_y:0.5
        outsideInput:  
# filter based on minimum rating
<RatingScreen>:
    BoxLayout:
        orientation: 'vertical'
        Label:
            text:"What is the minimum rated place, slider goes from 0-5"
            size_hint_y: 0.5
        Label:
            text:"Input 0 to skip, by moving input to far left"
            size_hint_y: 0.5
        Label:
            text:""
            size_hint_y:0.5
        ratingInput: 
#output 
<OutputScreen>:
    Label:
        text:""
        pos_hint_y: 0
        size_hint_x:0
        size_hint_y:0
    output: 

""")
#initialize worksheet
file = "fbla.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
ws = wb.active
#initialize lists of data
ListofAttraction = []
WebsiteofAttractions = []
for rows in ws.iter_rows(2):
    # add all 50 elements into data
    ListofAttraction.append(rows[0].value)
    WebsiteofAttractions.append(rows[1].value)
#class to add start button
class MainWidget(Widget):
    #initialize kivy class
    def __init__(self, **kwargs):
        super(MainWidget, self).__init__(**kwargs)

        #add button
        btnnext = Button(text='Start Program',pos=(0,500),size=(800,100))
        btnnext.bind(on_press=self.start)
        self.add_widget(btnnext)

    #add start function for button
    def start(self ,btn_inst):
        sm.current = "TypesSort"
#class for typeSort
class TypeInput(Widget):
    #initialize
    def __init__(self, **kwargs):
        super(TypeInput, self).__init__(**kwargs)
        #add new button for each type of sort
        btnnext = Button(text='Restaurant',pos=(0,100),size=(133,120))
        btnnext.bind(on_press=self.gonext)
        btnnext.bind(on_press=self.getText)
        self.add_widget(btnnext)
        btnnext2 = Button(text='zoos/theme parks', pos=(133, 100), size=(133, 120))
        btnnext2.bind(on_press=self.gonext)
        btnnext2.bind(on_press=self.getText2)
        self.add_widget(btnnext2)
        btnnext3 = Button(text='landmarks', pos=(133*2, 100), size=(133, 120))
        btnnext3.bind(on_press=self.gonext)
        btnnext3.bind(on_press=self.getText3)
        self.add_widget(btnnext3)
        btnnext4 = Button(text='areas', pos=(133*3, 100), size=(133, 120))
        btnnext4.bind(on_press=self.gonext)
        btnnext4.bind(on_press=self.getText4)
        self.add_widget(btnnext4)
        btnnext5 = Button(text='museums', pos=(133*4, 100), size=(133, 120))
        btnnext5.bind(on_press=self.gonext)
        btnnext5.bind(on_press=self.getText5)
        self.add_widget(btnnext5)
        btnnextX = Button(text='skip filter', pos=(133*5, 100), size=(133, 120))
        btnnextX.bind(on_press=self.gonext)
        btnnextX.bind(on_press=self.getTextX)
        self.add_widget(btnnextX)


    #get the text of button (there is no builtin functionality in kivy)
    def getText(self,btn_inst):
        global text
        text="restaurant"
    def getText2(self,btn_inst):
        global text
        text="zoos/theme parks"
    def getText3(self,btn_inst):
        global text
        text="landmark"
    def getText4(self,btn_inst):
        global text
        text="areas"
    def getText5(self,btn_inst):
        global text
        text="museum"
    def getTextX(self,btn_inst):
        global text
        text=""
    #actual filtering process
    def gonext(self ,btn_inst):
        #set new lists, that need to be accessed in different screen
        global FilteredList
        FilteredList=[]
        global FilteredWebsites
        FilteredWebsites = []
        if (FilteredList == []):
            #sets sorting value to TypesInput
            TypesInput = text
            #skip special case
            if (TypesInput == "none" or TypesInput == ""):
                FilteredList=ListofAttraction
                FilteredWebsites=WebsiteofAttractions
            #if type = TypesInput, add type to FilteredList
            for TypeRows in ws.iter_rows(2):
                if (TypeRows[2].value == TypesInput):
                    FilteredList.append(TypeRows[0].value)
                    FilteredWebsites.append(TypeRows[1].value)
        #check if FilteredList is no longer empty
        if(FilteredList!=[]):
            # move to new screen, resize window, and end program
            sm.current="AreaSort"
            Window.size = (800, 600)
            return FilteredList, FilteredWebsites
#class for areaSort
class areaInput(Widget):
    # initialize
    def __init__(self, **kwargs):
        super(areaInput, self).__init__(**kwargs)
        #add new button for each type of sort
        btnnext = Button(text='eastern', pos=(0, 100), size=(117, 120))
        btnnext.bind(on_press=self.gonext)
        btnnext.bind(on_press=self.getText)
        self.add_widget(btnnext)
        btnnext2 = Button(text='central', pos=(117, 100), size=(117, 120))
        btnnext2.bind(on_press=self.gonext)
        btnnext2.bind(on_press=self.getText2)
        self.add_widget(btnnext2)
        btnnext3 = Button(text='atlanta', pos=(117 * 2, 100), size=(133, 120))
        btnnext3.bind(on_press=self.gonext)
        btnnext3.bind(on_press=self.getText3)
        self.add_widget(btnnext3)
        btnnext4 = Button(text='coast', pos=(117 * 3, 100), size=(117, 120))
        btnnext4.bind(on_press=self.gonext)
        btnnext4.bind(on_press=self.getText4)
        self.add_widget(btnnext4)
        btnnext5 = Button(text='southern', pos=(117 * 4, 100), size=(117, 120))
        btnnext5.bind(on_press=self.gonext)
        btnnext5.bind(on_press=self.getText5)
        self.add_widget(btnnext5)
        btnnext6 = Button(text='mountains', pos=(117 * 5, 100), size=(117, 120))
        btnnext6.bind(on_press=self.gonext)
        btnnext6.bind(on_press=self.getText6)
        self.add_widget(btnnext6)
        btnnextX = Button(text='skip filter', pos=(117*6, 100), size=(117, 120))
        btnnextX.bind(on_press=self.gonext)
        btnnextX.bind(on_press=self.getTextX)
        # add output Label for special circumstances
        self.add_widget(btnnextX)
        self.areaText= Label(pos=(0,0),size=(800,120))
        self.add_widget(self.areaText)
    # get Text of all buttons
    def getText(self,btn_inst):
        global atext
        atext="eastern"
    def getText2(self,btn_inst):
        global atext
        atext="central"
    def getText3(self,btn_inst):
        global atext
        atext="atlanta"
    def getText4(self,btn_inst):
        global atext
        atext="coast"
    def getText5(self,btn_inst):
        global atext
        atext="southern"
    def getText6(self,btn_inst):
        global atext
        atext="mountains"
    def getTextX(self,btn_inst):
        global atext
        atext=""
    def gonext(self ,btn_inst):
        # make new lists
        global AreaFilteredList
        AreaFilteredList = []
        global AreaFilteredWebsites
        AreaFilteredWebsites = []
        if (AreaFilteredList == []):
            #set sorting variable
            AreaInput = atext
            #special case for skipped
            if (AreaInput == "none" or AreaInput == ""):
                AreaFilteredList = FilteredList
                AreaFilteredWebsites = FilteredWebsites
            for AreaRows in ws.iter_rows(2):
                #add values to list
                if ((AreaRows[3].value) == AreaInput and (AreaRows[0].value) in FilteredList):
                    AreaFilteredList.append(AreaRows[0].value)
                    AreaFilteredWebsites.append(AreaRows[1].value)
            if (AreaFilteredList == []):
                #special case if there are no locations that fit filters
                self.txt2 = "Oops, there are no choices for your filters"
                self.areaText.text = self.txt2
            #finish function
            if (AreaFilteredList != []):
                sm.current = "SurroundSort"
                Window.size=(800,600)
                return AreaFilteredList, AreaFilteredWebsites
#surroundsort
class surroundInput(Widget):
    #initialize
    def __init__(self, **kwargs):
        super(surroundInput, self).__init__(**kwargs)
        print("\nmainwidget:")
        #make all labels
        btnnext = Button(text='urban', pos=(0, 100), size=(200, 120))
        btnnext.bind(on_press=self.gonext)
        btnnext.bind(on_press=self.getText)
        self.add_widget(btnnext)
        btnnext2 = Button(text='suburban', pos=(200, 100), size=(200, 120))
        btnnext2.bind(on_press=self.gonext)
        btnnext2.bind(on_press=self.getText2)
        self.add_widget(btnnext2)
        btnnext3 = Button(text='rural', pos=(400, 100), size=(200, 120))
        btnnext3.bind(on_press=self.gonext)
        btnnext3.bind(on_press=self.getText3)
        self.add_widget(btnnext3)
        btnnextX = Button(text='skip filter', pos=(600, 100), size=(200, 120))
        btnnextX.bind(on_press=self.gonext)
        btnnextX.bind(on_press=self.getTextX)
        #label for special cases
        self.add_widget(btnnextX)
        self.surroundText= Label(pos=(0,0),size=(800,120))
        self.add_widget(self.surroundText)
    #gettext
    def getText(self,btn_inst):
        global stext
        stext="urban"
    def getText2(self,btn_inst):
        global stext
        stext="suburban"
    def getText3(self,btn_inst):
        global stext
        stext="rural"
    def getTextX(self,btn_inst):
        global stext
        stext=""
    # function for sorting
    def gonext(self,btn_inst):
        #initialize new lists
        global SurroundFilteredList
        SurroundFilteredList = []
        global SurroundFilteredWebsites
        SurroundFilteredWebsites = []
        if (SurroundFilteredList == []):
            #set sort variable
            SouroundInput = stext
            #special case: skip button
            if (SouroundInput == "none" or SouroundInput == ""):
                SurroundFilteredList = AreaFilteredList
                SurroundFilteredWebsites = AreaFilteredWebsites
            # sort function
            for SurroundRows in ws.iter_rows(2):
                if ((SurroundRows[5].value) == SouroundInput and (SurroundRows[0].value) in AreaFilteredList):
                    SurroundFilteredList.append(SurroundRows[0].value)
                    SurroundFilteredWebsites.append(SurroundRows[1].value)
            if (SurroundFilteredList == []):
                #special case: no locations that fir filters
                self.txt2 = "Oops, there are no choices for your filters"
                self.surroundText.text = self.txt2
            # end function
            if (SurroundFilteredList != []):
                sm.current = "OutsideSort"
                Window.size = (800, 600)
                return SurroundFilteredList, SurroundFilteredWebsites
#inside or outside location
class outsideInput(Widget):
    #initialize
    def __init__(self, **kwargs):
        super(outsideInput, self).__init__(**kwargs)
        # add yes, no and skip buttons
        btnnext = Button(text='yes', pos=(0, 100), size=(267, 120))
        btnnext.bind(on_press=self.gonext)
        btnnext.bind(on_press=self.getText)
        self.add_widget(btnnext)
        btnnext2 = Button(text='no', pos=(267, 100), size=(267, 120))
        btnnext2.bind(on_press=self.gonext)
        btnnext2.bind(on_press=self.getText2)
        self.add_widget(btnnext2)
        btnnextX = Button(text='skip filter', pos=(533, 100), size=(267, 120))
        btnnextX.bind(on_press=self.gonext)
        btnnextX.bind(on_press=self.getTextX)
        self.add_widget(btnnextX)
        self.outsideText= Label(pos=(0,0),size=(800,120))
        self.add_widget(self.outsideText)
    # get yes, no, or ""
    def getText(self,btn_inst):
        global otext
        otext="yes"
    def getText2(self,btn_inst):
        global otext
        otext="no"
    def getTextX(self,btn_inst):
        global otext
        otext=""
    #sort
    def gonext(self,btn_inst):
        #initialize list
        global OutsideFilteredList
        OutsideFilteredList = []
        global OutsideFilteredWebsites
        OutsideFilteredWebsites = []
        #sort
        if (OutsideFilteredList == []):
            OutsideInput = otext
            #set sort variable^
            #special case: skipv
            if (OutsideInput == "none" or OutsideInput == ""):
                OutsideFilteredList = SurroundFilteredList
                OutsideFilteredWebsites = SurroundFilteredWebsites
            #actual filter
            for OutsideRows in ws.iter_rows(2):
                if ((OutsideRows[4].value) == OutsideInput and (OutsideRows[0].value) in SurroundFilteredList):
                    OutsideFilteredList.append(OutsideRows[0].value)
                    OutsideFilteredWebsites.append(OutsideRows[1].value)
            if (OutsideFilteredList == []):
                #special case: no items in list
                self.txt2 = "Oops, there are no choices for your filters"
                self.outsideText.text =self.txt2
            #end program
            if (OutsideFilteredList != []):
                sm.current = "RatingSort"
                Window.size=(800,600)
                return OutsideFilteredList, OutsideFilteredWebsites
#ratingSort
class ratingInput(Widget):
    #set value of slider as integer
    slider_val = NumericProperty(0)
    #initialize
    def __init__(self, **kwargs):
        super(ratingInput, self).__init__(**kwargs)
        #finish button
        btnnext = Button(text='Finish',pos=(0,0),size=(267,120))
        btnnext.bind(on_press=self.gonext)
        self.add_widget(btnnext)
        #label for special cases
        self.ratingText=Label(pos=(267,0),size=(267,120))
        self.add_widget(self.ratingText)
        #add slider that takes input
        self.slider=Slider(min=0,max=5,pos=(0,120),size=(800,120))
        self.slider.fbind('value',self.on_slider_val)
        self.add_widget(self.slider)
        #set label that displays current value of slider
        self.dispNum=Label(pos=(533,0),size=(267,120))
        self.add_widget(self.dispNum)
    #Function to display current value of slider
    def on_slider_val(self, instance, val):
        self.dispNum.text = str(val)

    #sort
    def gonext(self,btn_inst):
        #make lists
        global FinalList
        FinalList = []
        global FinalWebsites
        FinalWebsites = []
        #sets minimum value that you want attraction to be rated
        MinimumRating =self.slider.value
        #sort
        for RatingRows in ws.iter_rows(2):
            if (float((RatingRows[6].value)) >= MinimumRating and (RatingRows[0].value) in OutsideFilteredList):
                FinalList.append(RatingRows[0].value)
                FinalWebsites.append(RatingRows[1].value)
        #special case: no items in FinalList
        if (FinalList == []):

            self.txt2 = "Oops, there are no choices for your filters"
            self.ratingText.text = self.txt2
        #end function
        if (FinalList != []):
            sm.current = "OutputScreen"
            Window.size=(800,600)
            return FinalList, FinalWebsites
#output
class output(Widget):
    #initialize
    def __init__(self, **kwargs):
        super(output, self).__init__(**kwargs)
        #build/rebuild screen
        btnnext = Button(text='Build - After restart must build again to wipe', pos=(0,550 ), size=(400,50))
        btnnext.bind(on_press=self.build)
        self.add_widget(btnnext)
        #restart program
        restart = Button(text='Restart', pos=(400, 550), size=(400, 50))
        restart.bind(on_press=self.wipe)
        self.add_widget(restart)
    #build screen


    def build(self,btn_inst):
        global FinalWebz
        if (len(FinalList) > 25):
            FinalList2=FinalList[0:25]
            FinalWebz=FinalWebsites[0:25]
        else:
            FinalList2=FinalList
            FinalWebz=FinalWebsites
        for i in range(len(FinalList2)):
            #get size of all items
            space=550/len(FinalList2)
            #dynamically change fonts when dealing with large sizes
            if(len(FinalList2)>24):
                r=500/len(FinalList2)
            else:
                r=20

            # build each button for attraction
            self.webs=Button(text=FinalList2[i]+" id:"+str(i),size=(800,space),pos=(000,space*i),font_size=r)
            self.webs.bind(on_press=self.website)
            self.add_widget(self.webs)
    #click to go to website
    def website(self,btn_inst):
        key=btn_inst.text.split(":")
        webbrowser.open(FinalWebz[int(key[1])])
    #wipe all data
    def wipe(self,btn_inst):
        FilteredList=[]
        FilteredWebsites=[]
        AreaFilteredList=[]
        AreaFilteredWebsites=[]
        SurroundFilteredList=[]
        SurroundFilteredWebsites=[]
        OutsideFilteredList=[]
        OutsideFilteredWebsites=[]
        FinalList=[]
        FinalWebsites=[]
        sm.current="menu"
        Window.size = (800, 600)






#create all screens
class startScreen(Screen):
    pass


class typesScreen(Screen):
    pass
class areasScreen(Screen):
    pass
class SurroundScreen(Screen):
    pass
class OutsideScreen(Screen):
    pass
class RatingScreen(Screen):
    pass
class OutputScreen(Screen):
    pass
class ScreenManager(ScreenManager):
    pass

# Create the screen manager
sm = ScreenManager()
sm.add_widget(startScreen(name='menu'))
sm.add_widget(typesScreen(name='TypesSort'))
sm.add_widget(areasScreen(name='AreaSort'))
sm.add_widget(SurroundScreen(name='SurroundSort'))
sm.add_widget(OutsideScreen(name="OutsideSort"))
sm.add_widget(RatingScreen(name="RatingSort"))
sm.add_widget(OutputScreen(name="OutputScreen"))
class TestApp(App):
    #build app
    def build(self):
        return sm

#run app
if __name__ == '__main__':
    TestApp().run()